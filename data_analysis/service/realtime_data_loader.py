import logging
import logging.config
import threading
import time
from datetime import datetime, timedelta

import openpyxl
#from analysis.search_trends_analysis.analysis import TimeSeriesAnalysis
from meteostat import Daily, Stations
from data_analysis.loader.module.analyze import DataAnalyzer
from data_analysis.loader.module.loader import DataLoader


class RealtimeMonitoringService:
    def __init__(self, service_id, kw_list, timeframe, config_path='configs/logging.conf'):
        self.service_id = service_id
        self.kw_list = kw_list
        self.timeframe = timeframe
        self.stop_event = threading.Event()
        self.thread = None
        self.loader = DataLoader([49.2497, -123.1193])
        logging.config.fileConfig(config_path)
        self.logger = logging.getLogger(f"service_{self.service_id}")

    def set_coords(self, coords):
        self.loader.set_coords(coords)

    def start(self):
        if self.thread is None or not self.thread.is_alive():
            self.logger.info(f"Service {self.service_id} starting")
            self.thread = threading.Thread(target=self.run, daemon=True)
            self.thread.start()
            self.logger.info(f"Service {self.service_id} started successfully")

    def stop(self):
        if self.thread is not None:
            self.stop_event.set()
            self.logger.info(f"Service {self.service_id} stopping")
            self.thread.join(timeout=5)
            if self.thread.is_alive():
                self.logger.warning(f"Service {self.service_id} did not stop gracefully")
            self.thread = None
            self.logger.info(f"Service {self.service_id} stopped")

    def run(self):
        while not self.stop_event.is_set():
            try:
                self.load_and_analyze()
            except Exception as e:
                self.logger.exception(f"Service {self.service_id} encountered an error: {e}")
            time.sleep(5)

    def load_and_analyze(self):
        # Get daily data for 2018
        self.loader.refresh()
        data = self.loader.get_data()
        pymeteo = DataAnalyzer(data)
        try:
            data = pymeteo.get_data()

            if data is not None and not data.empty:
                pass
                data = data.infer_objects(copy=False)
                data.reset_index(inplace=True)
                print(self.coords)
                

                #Analisis
                #pymeteo.moving_average(2)
                pymeteo.calculate_difference(1)
                pymeteo.autocorrelation(1, 2)
                pymeteo.find_max(3)
                pymeteo.find_min(3)
                # self.save_result_to_file(pymeteo.get_data())
                #self.append_dataframe_to_excel(pymeteo.get_data(), 'data.xlsx', 'Sheet1')
               # pymeteo.get_data().to_excel('data.xlsx', index=False)

                self.logger.info(f"Service {self.service_id}: Analysis results saved to file.")
            else:
                self.logger.warning(f"Service {self.service_id}: No data received from Meteostat.")
        except Exception as e:
            self.logger.exception(f"Service {self.service_id}: Error during data loading: {e}")

    import pandas as pd

    def append_dataframe_to_excel(self, df: pd.DataFrame, filename: str, sheet_name: str = 'Sheet1') -> None:
        """
        Добавляет DataFrame в существующий Excel-файл или создает новый.

        Args:
        df (pd.DataFrame): DataFrame для добавления в Excel.
        filename (str): Имя файла Excel.
        sheet_name (str): Имя листа в Excel (по умолчанию 'Sheet1').
        """
        try:
            # Проверяем, существует ли файл
            workbook = openpyxl.load_workbook(filename=filename)

            # Если лист с заданным именем не существует, создаем его
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)

            # Выбираем лист по имени
            worksheet = workbook[sheet_name]

            # Добавляем данные в лист
            for row_idx, row in df.iterrows():
                worksheet.append(row.values)

            # Сохраняем изменения
            workbook.save(filename)
            print(f"Данные успешно добавлены в файл {filename}")
        except Exception as e:
            print(f"Ошибка при добавлении данных в файл {filename}: {str(e)}")

    def save_result_to_file(self, result: pd.DataFrame) -> None:
        """Сохраняет результаты анализа в текстовый файл."""
        with open(f"./analysis_results.txt", "a") as file:
            file.write(f"Result:\n{result}\n")
            file.write(f"{'-' * 40}\n")
